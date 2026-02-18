from __future__ import annotations

import csv
import os
from dataclasses import asdict
from pathlib import Path
from typing import List, Optional, Tuple

import click

from netops.security.passwords import (
    PasswordPolicy,
    generate_password,
    parse_format,
    DEFAULT_SYMBOLS,
    DIGITS_0_9,
    DIGITS_1_9,
)

# XLSX support is optional; only imported if used
def _load_openpyxl():
    try:
        import openpyxl  # type: ignore
        return openpyxl
    except Exception as e:
        raise click.ClickException(
            "openpyxl is required for XLSX support. Install it (pip install openpyxl)."
        ) from e


def _infer_out_kind(out_path: Optional[Path], out_kind: str) -> str:
    if out_kind != "auto":
        return out_kind
    if not out_path:
        return "screen"
    ext = out_path.suffix.lower()
    if ext in (".txt",):
        return "txt"
    if ext in (".csv",):
        return "csv"
    if ext in (".xlsx", ".xlsm"):
        return "xlsx"
    # default if unknown extension
    return "txt"


def _normalize_fmt(fmt: str, length: Optional[int]) -> str:
    """
    Your helper generates exactly len(tokens) when fmt is provided.
    For CLI convenience:
      - if length is None: use fmt as-is
      - if length is provided: repeat fmt pattern to reach length, then truncate
    """
    tokens = parse_format(fmt)
    if length is None:
        return "".join(tokens)
    if length <= 0:
        raise click.ClickException("--length must be > 0.")
    rep = (tokens * ((length + len(tokens) - 1) // len(tokens)))[:length]
    return "".join(rep)


def _render_table(passwords: List[str]) -> str:
    # simple "dataframe-ish" table without external deps
    idx_w = max(len(str(len(passwords))), 1)
    pw_w = max((len(p) for p in passwords), default=8)
    sep = f"+-{'-'*idx_w}-+-{'-'*pw_w}-+"
    out = [sep, f"| {'#'.rjust(idx_w)} | {'password'.ljust(pw_w)} |", sep]
    for i, pw in enumerate(passwords, start=1):
        out.append(f"| {str(i).rjust(idx_w)} | {pw.ljust(pw_w)} |")
    out.append(sep)
    return "\n".join(out)


def _write_txt(path: Path, passwords: List[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(passwords) + "\n", encoding="utf-8")


def _write_csv(path: Path, passwords: List[str], header: str = "password") -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=[header])
        w.writeheader()
        for pw in passwords:
            w.writerow({header: pw})


def _write_xlsx(path: Path, passwords: List[str]) -> None:
    import pandas as pd
    from netops.excel import write_workbook  # your helper

    path.parent.mkdir(parents=True, exist_ok=True)

    df = pd.DataFrame(
        {
            "index": list(range(1, len(passwords) + 1)),
            "password": passwords,
        }
    )

    # write_workbook expects: list[tuple[prop, system, df]]
    results = [("Passwords", "pw-gen", df)]
    write_workbook(str(path), results)



def _read_csv_rows(path: Path) -> Tuple[List[dict], List[str]]:
    with path.open("r", newline="", encoding="utf-8") as f:
        r = csv.DictReader(f)
        rows = list(r)
        fieldnames = list(r.fieldnames or [])
    return rows, fieldnames


def _write_csv_rows(path: Path, rows: List[dict], fieldnames: List[str]) -> None:
    tmp = path.with_suffix(path.suffix + ".tmp")
    with tmp.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow({k: row.get(k, "") for k in fieldnames})
    tmp.replace(path)


def _append_passwords_to_csv(
    file_path: Path,
    passwords: List[str],
    column: str = "password",
    fill_only_blanks: bool = True,
) -> int:
    rows, fieldnames = _read_csv_rows(file_path)
    if not rows:
        # empty file with only header or nothing
        if column not in fieldnames:
            fieldnames = fieldnames + [column] if fieldnames else [column]
        # create rows from passwords
        new_rows = [{column: pw} for pw in passwords]
        _write_csv_rows(file_path, new_rows, fieldnames)
        return len(passwords)

    if column not in fieldnames:
        fieldnames.append(column)

    used = 0
    pw_iter = iter(passwords)
    for row in rows:
        if used >= len(passwords):
            break
        current = (row.get(column) or "").strip()
        if fill_only_blanks and current:
            continue
        row[column] = next(pw_iter)
        used += 1

    # if passwords remain after filling blanks, append new rows
    remaining = passwords[used:]
    for pw in remaining:
        rows.append({**{k: "" for k in fieldnames}, column: pw})
        used += 1

    _write_csv_rows(file_path, rows, fieldnames)
    return used


def _append_passwords_to_xlsx(
    file_path: Path,
    passwords: List[str],
    sheet: Optional[str],
    column: str = "password",
    header_row: int = 1,
    fill_only_blanks: bool = True,
) -> int:
    openpyxl = _load_openpyxl()

    wb = openpyxl.load_workbook(file_path)
    ws = wb[sheet] if sheet else wb.active

    # Find or create column by header name in header_row
    max_col = ws.max_column or 1
    col_idx = None
    for c in range(1, max_col + 1):
        v = ws.cell(row=header_row, column=c).value
        if isinstance(v, str) and v.strip().lower() == column.lower():
            col_idx = c
            break
    if col_idx is None:
        col_idx = max_col + 1
        ws.cell(row=header_row, column=col_idx, value=column)

    used = 0
    pw_i = 0
    # Fill blanks in existing rows first (starting after header)
    for r in range(header_row + 1, (ws.max_row or header_row) + 1):
        if pw_i >= len(passwords):
            break
        current = ws.cell(row=r, column=col_idx).value
        current_s = str(current).strip() if current is not None else ""
        if fill_only_blanks and current_s:
            continue
        ws.cell(row=r, column=col_idx, value=passwords[pw_i])
        pw_i += 1
        used += 1

    # Append any remaining passwords as new rows
    r = (ws.max_row or header_row) + 1
    while pw_i < len(passwords):
        ws.cell(row=r, column=col_idx, value=passwords[pw_i])
        pw_i += 1
        used += 1
        r += 1

    wb.save(file_path)
    return used


@click.command("pw-gen")
@click.option(
    "-f",
    "--format",
    "fmt",
    default=None,
    help='Format tokens (u/l/n/s). Examples: "ulnnnnns" or "u,l,n,n,n,n,n,s". If set, pattern is used (optionally repeated to --length).',
)
@click.option(
    "-l",
    "--length",
    type=int,
    default=8,
    show_default=True,
    help="Password length (ignored if --format is set unless the format is repeated/truncated to this length).",
)
@click.option(
    "-n",
    "--count",
    "--amount",
    type=int,
    default=1,
    show_default=True,
    help="How many passwords to generate.",
)
@click.option("--no-upper", is_flag=True, help="Disable uppercase letters for random-policy generation.")
@click.option("--no-lower", is_flag=True, help="Disable lowercase letters for random-policy generation.")
@click.option("--no-digits", is_flag=True, help="Disable digits for random-policy generation.")
@click.option("--no-symbols", is_flag=True, help="Disable symbols for random-policy generation.")
@click.option("--symbols", default=DEFAULT_SYMBOLS, show_default=True, help="Symbols set to use when enabled.")
@click.option(
    "--digits",
    type=click.Choice(["0-9", "1-9"], case_sensitive=False),
    default="0-9",
    show_default=True,
    help="Digit set to use when digits are enabled.",
)
@click.option(
    "-o",
    "--output",
    "out_path",
    type=click.Path(dir_okay=False, writable=True, path_type=Path),
    default=None,
    help="Write generated passwords to a file. If omitted, prints to screen.",
)
@click.option(
    "--out-kind",
    type=click.Choice(["auto", "screen", "txt", "csv", "xlsx"], case_sensitive=False),
    default="auto",
    show_default=True,
    help="Output format. 'auto' infers from --output extension.",
)
@click.option(
    "--append-to",
    "append_path",
    type=click.Path(exists=True, dir_okay=False, path_type=Path),
    default=None,
    help="Load an existing CSV/XLSX and add/fill a password column.",
)
@click.option(
    "--append-sheet",
    default=None,
    help="(XLSX) Sheet name to modify. Default: active sheet.",
)
@click.option(
    "--append-column",
    default="Password",
    show_default=True,
    help="Column name to create/fill when using --append-to.",
)
@click.option(
    "--fill-only-blanks/--overwrite-existing",
    default=True,
    show_default=True,
    help="When appending, fill only blank cells/fields or overwrite existing values.",
)
@click.option(
    "--header-row",
    type=int,
    default=1,
    show_default=True,
    help="(XLSX) Header row index.",
)
def pw_gen_cli(
    fmt: Optional[str],
    length: int,
    count: int,
    no_upper: bool,
    no_lower: bool,
    no_digits: bool,
    no_symbols: bool,
    symbols: str,
    digits: str,
    out_path: Optional[Path],
    out_kind: str,
    append_path: Optional[Path],
    append_sheet: Optional[str],
    append_column: str,
    fill_only_blanks: bool,
    header_row: int,
) -> None:
    """
    Generate passwords using netops.security.passwords.generate_password().
    """

    if count <= 0:
        raise click.ClickException("--count/--amount must be > 0.")
    if length <= 0:
        raise click.ClickException("--length must be > 0.")

    digset = DIGITS_0_9 if digits.lower() == "0-9" else DIGITS_1_9

    # If we're appending to a file, and user left count at default 1,
    # it's usually because they want "enough to fill rows". We'll try to infer.
    if append_path is not None and count == 1:
        ext = append_path.suffix.lower()
        if ext == ".csv":
            rows, _ = _read_csv_rows(append_path)
            if rows:
                # count = number of blanks (or all rows if overwrite)
                blanks = 0
                for r in rows:
                    v = (r.get(append_column) or "").strip()
                    if (not v) or (not fill_only_blanks):
                        blanks += 1
                count = max(blanks, 1)
        elif ext in (".xlsx", ".xlsm"):
            openpyxl = _load_openpyxl()
            wb = openpyxl.load_workbook(append_path)
            ws = wb[append_sheet] if append_sheet else wb.active

            # Find column index (if exists)
            col_idx = None
            for c in range(1, (ws.max_column or 1) + 1):
                v = ws.cell(row=header_row, column=c).value
                if isinstance(v, str) and v.strip().lower() == append_column.lower():
                    col_idx = c
                    break
            blanks = 0
            if ws.max_row and ws.max_row > header_row:
                for r in range(header_row + 1, ws.max_row + 1):
                    if col_idx is None:
                        blanks += 1  # column doesn't exist yet; treat all as blank
                        continue
                    v = ws.cell(row=r, column=col_idx).value
                    v_s = str(v).strip() if v is not None else ""
                    if (not v_s) or (not fill_only_blanks):
                        blanks += 1
            count = max(blanks, 1)

    # Build policy for random-policy generation (fmt=None)
    policy = PasswordPolicy(
        length=length,
        use_upper=not no_upper,
        use_lower=not no_lower,
        use_digits=not no_digits,
        use_symbols=not no_symbols,
        symbols=symbols,
        digits=digset,
    )

    # Normalize fmt behavior
    fmt_final = None
    if fmt:
        fmt_final = _normalize_fmt(fmt, length)

    passwords: List[str] = []
    for _ in range(count):
        passwords.append(
            generate_password(
                policy=policy,
                fmt=fmt_final,
                symbols=symbols,
                digits=digset,
            )
        )

    # If appending to an existing file, do that first (and optionally also output)
    if append_path is not None:
        ext = append_path.suffix.lower()
        if ext == ".csv":
            used = _append_passwords_to_csv(
                append_path,
                passwords,
                column=append_column,
                fill_only_blanks=fill_only_blanks,
            )
            click.echo(f"Updated CSV: {append_path} (wrote {used} passwords to column '{append_column}').")
        elif ext in (".xlsx", ".xlsm"):
            used = _append_passwords_to_xlsx(
                append_path,
                passwords,
                sheet=append_sheet,
                column=append_column,
                header_row=header_row,
                fill_only_blanks=fill_only_blanks,
            )
            click.echo(f"Updated XLSX: {append_path} (wrote {used} passwords to column '{append_column}').")
        else:
            raise click.ClickException("--append-to only supports .csv, .xlsx, .xlsm")

    # Output
    kind = _infer_out_kind(out_path, out_kind.lower())
    if kind == "screen" or out_path is None:
        if len(passwords) == 1:
            click.echo(passwords[0])
        else:
            click.echo(_render_table(passwords))
        return

    if kind == "txt":
        _write_txt(out_path, passwords)
    elif kind == "csv":
        _write_csv(out_path, passwords, header="password")
    elif kind == "xlsx":
        _write_xlsx(out_path, passwords, sheet="passwords", header="password")
    else:
        raise click.ClickException(f"Unsupported out kind: {kind}")

    click.echo(f"Wrote {len(passwords)} passwords -> {out_path}")
